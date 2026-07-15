from fastapi import FastAPI, UploadFile, File, Form, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
import shutil
import os
import json
import re
import io
import asyncio
from services.ocr_service import OCRService
from services.eval_service import EvalService
from services.db_service import init_db, insert_result, get_all_results, get_result_by_roll, delete_result
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, TableStyle, Paragraph, Spacer
from reportlab.lib.units import inch
from datetime import datetime

load_dotenv()

app = FastAPI(title="ScriptSense API", version="2.0")
init_db()

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# ✅ NEW: 3 টে আলাদা Upload Folder
UPLOAD_BASE = "uploads"
QP_FOLDER = os.path.join(UPLOAD_BASE, "question_papers")
MODEL_FOLDER = os.path.join(UPLOAD_BASE, "model_answers")
SCRIPT_FOLDER = os.path.join(UPLOAD_BASE, "student_scripts")

os.makedirs(QP_FOLDER, exist_ok=True)
os.makedirs(MODEL_FOLDER, exist_ok=True)
os.makedirs(SCRIPT_FOLDER, exist_ok=True)
os.makedirs("cache", exist_ok=True) # image_paths এর জন্য রাখলাম
os.makedirs("results", exist_ok=True)

app.mount("/cache", StaticFiles(directory="cache"), name="cache")

ocr_service = OCRService()
eval_service = EvalService()

def sanitize_filename(filename):
    return re.sub(r'[^\w\-_\.]', '_', filename)

def get_timestamp():
    return datetime.now().strftime('%Y%m%d_%H%M%S')

# ==================== SSE EVALUATION STREAM ====================

@app.post("/evaluate-stream")
async def evaluate_stream_endpoint(
    script: UploadFile = File(...),
    question_paper: UploadFile = File(...),
    model_answer: UploadFile = File(None),
    model_text: str = Form(None),
    max_marks: int = Form(25)
):
    async def evaluation_stream():
        try:
            # ✅ CHANGED: আলাদা Folder + Timestamp
            student_path = f"{SCRIPT_FOLDER}/SCRIPT_{get_timestamp()}_{sanitize_filename(script.filename)}"
            qp_path = f"{QP_FOLDER}/QP_{get_timestamp()}_{sanitize_filename(question_paper.filename)}"
            model_path = None

            with open(student_path, "wb") as f:
                shutil.copyfileobj(script.file, f)
            with open(qp_path, "wb") as f:
                shutil.copyfileobj(question_paper.file, f)

            if model_answer:
                model_path = f"{MODEL_FOLDER}/MODEL_{get_timestamp()}_{sanitize_filename(model_answer.filename)}"
                with open(model_path, "wb") as f:
                    shutil.copyfileobj(model_answer.file, f)

            yield f"data: {json.dumps({'progress': 10, 'status': 'Question Paper Uploaded', 'complete': False})}\n\n"
            await asyncio.sleep(0.2)

            yield f"data: {json.dumps({'progress': 20, 'status': 'Model Answer Uploaded', 'complete': False})}\n\n"
            await asyncio.sleep(0.2)

            yield f"data: {json.dumps({'progress': 30, 'status': 'Student Script Uploaded', 'complete': False})}\n\n"
            await asyncio.sleep(0.2)

            yield f"data: {json.dumps({'progress': 40, 'status': 'Extracting Text', 'complete': False})}\n\n"
            student_data = ocr_service.process_pdf(student_path)
            student_text = student_data['text']
            bboxes = student_data.get('bboxes', [])
            image_paths = student_data.get('image_paths', [])
            roll = student_data.get('roll_no', 'NOT_FOUND')
            name = student_data.get('name', 'NOT_FOUND')
            await asyncio.sleep(0.3)

            yield f"data: {json.dumps({'progress': 50, 'status': 'Cleaning Text', 'complete': False})}\n\n"
            qp_text = ocr_service.process_pdf(qp_path)['text']
            model_text_extracted = ocr_service.process_pdf(model_path)['text'] if model_path else model_text or ""
            await asyncio.sleep(0.3)

            yield f"data: {json.dumps({'progress': 60, 'status': 'Question Segmentation', 'complete': False})}\n\n"
            await asyncio.sleep(0.3)

            yield f"data: {json.dumps({'progress': 70, 'status': 'Dynamic Rubric Generation', 'complete': False})}\n\n"
            qp_rubric, _ = eval_service.extract_rubric_from_qp(qp_text)
            model_answers = eval_service.extract_answers_from_model(
                model_text_extracted,
                qp_rubric.keys(),
                qp_rubric
            )
            await asyncio.sleep(0.3)

            yield f"data: {json.dumps({'progress': 80, 'status': 'Semantic Evaluation', 'complete': False})}\n\n"
            result = eval_service.evaluate(student_text, model_text_extracted, qp_rubric, bboxes)
            await asyncio.sleep(0.3)

            yield f"data: {json.dumps({'progress': 90, 'status': 'Score Calculation', 'complete': False})}\n\n"

            subject_code = "NOT_FOUND"
            subject_name = "NOT_FOUND"
            code_match = re.search(r'(?:Subject|Course|Paper)\s*Code[\s\:\-]*([A-Z]{2,6}[-]?\d{2,4})', qp_text, re.I)
            if code_match:
                subject_code = code_match.group(1).strip().upper()

            name_match = re.search(r'(?:Subject|Course|Paper)[\s\:\-]*(?:Name)?[\s\:\-]*([A-Z][A-Za-z\s&,\-]{3,60})', qp_text, re.I)
            if name_match:
                subject_name = name_match.group(1).strip()
                subject_name = re.sub(r'\s+(Code|Full|Marks|Total).*$', '', subject_name, flags=re.I).strip()

            if roll not in ['ERROR', ''] and subject_name not in ['ERROR', '']:
                insert_result(
                    roll_no=roll,
                    subject_code=subject_code,
                    subject_name=subject_name,
                    name=name,
                    max_marks=result['total_max'],
                    marks_obtained=result['total_score'],
                    criterion=result['feedback'],
                    feedback=result['feedback'],
                    teacher_feedback="",
                    image_paths=image_paths,
                    bboxes=[],
                    markings=result['markings']
                )
            await asyncio.sleep(0.2)

            yield f"data: {json.dumps({'progress': 100, 'status': 'Feedback Generation', 'complete': True, 'roll_no': roll})}\n\n"

        except Exception as e:
            yield f"data: {json.dumps({'progress': 0, 'status': f'Error: {str(e)}', 'complete': False, 'error': True})}\n\n"

    return StreamingResponse(evaluation_stream(), media_type="text/event-stream")

@app.post("/evaluate-batch-stream")
async def evaluate_batch_stream_endpoint(
    scripts: list[UploadFile] = File(...),
    question_paper: UploadFile = File(None),
    model_answer: UploadFile = File(None),
    model_text: str = Form(None),
    max_marks: int = Form(25)
):
    async def batch_evaluation_stream():
        try:
            total = len(scripts)
            qp_path = None
            model_path = None
            qp_text = ""
            model_text_extracted = model_text or ""

            if question_paper:
                qp_path = f"{QP_FOLDER}/QP_{get_timestamp()}_{sanitize_filename(question_paper.filename)}"
                with open(qp_path, "wb") as f:
                    shutil.copyfileobj(question_paper.file, f)
                qp_text = ocr_service.process_pdf(qp_path)['text']

            if model_answer:
                model_path = f"{MODEL_FOLDER}/MODEL_{get_timestamp()}_{sanitize_filename(model_answer.filename)}"
                with open(model_path, "wb") as f:
                    shutil.copyfileobj(model_answer.file, f)
                model_text_extracted = ocr_service.process_pdf(model_path)['text']

            qp_rubric, _ = eval_service.extract_rubric_from_qp(qp_text) if qp_text else ({}, {})
            model_answers = eval_service.extract_answers_from_model(
                model_text_extracted,
                qp_rubric.keys(),
                qp_rubric
            )

            for idx, script in enumerate(scripts):
                base_progress = int((idx / total) * 90)
                yield f"data: {json.dumps({'progress': base_progress, 'status': f'Processing {script.filename} ({idx+1}/{total})', 'complete': False})}\n\n"

                script_path = f"{SCRIPT_FOLDER}/SCRIPT_{get_timestamp()}_{sanitize_filename(script.filename)}"
                with open(script_path, "wb") as f:
                    shutil.copyfileobj(script.file, f)

                student_data = ocr_service.process_pdf(script_path)
                student_text = student_data['text']
                bboxes = student_data.get('bboxes', [])
                image_paths = student_data.get('image_paths', [])
                roll = student_data.get('roll_no', 'NOT_FOUND')
                name = student_data.get('name', 'NOT_FOUND')

                result = eval_service.evaluate(
        student_text,
        model_text_extracted,
        qp_rubric,
        bboxes,
        student_images=image_paths # ✅ eta add koro
)
                subject_code = "NOT_FOUND"
                subject_name = "NOT_FOUND"
                if qp_text:
                    code_match = re.search(r'(?:Subject|Course|Paper)\s*Code[\s\:\-]*([A-Z]{2,6}[-]?\d{2,4})', qp_text, re.I)
                    if code_match:
                        subject_code = code_match.group(1).strip().upper()
                    name_match = re.search(r'(?:Subject|Course|Paper)[\s\:\-]*(?:Name)?[\s\:\-]*([A-Z][A-Za-z\s&,\-]{3,60})', qp_text, re.I)
                    if name_match:
                        subject_name = name_match.group(1).strip()
                        subject_name = re.sub(r'\s+(Code|Full|Marks|Total).*$', '', subject_name, flags=re.I).strip()

                if roll not in ['ERROR', ''] and subject_name not in ['ERROR', '']:
                    insert_result(
                        roll_no=roll,
                        subject_code=subject_code,
                        subject_name=subject_name,
                        name=name,
                        max_marks=result['total_max'],
                        marks_obtained=result['total_score'],
                        criterion=result['feedback'],
                        feedback=result['feedback'],
                        teacher_feedback="",
                        image_paths=image_paths,
                        bboxes=[],
                        markings=result['markings']
                    )

                await asyncio.sleep(0.3)

            yield f"data: {json.dumps({'progress': 100, 'status': 'Batch Complete', 'complete': True})}\n\n"

        except Exception as e:
            yield f"data: {json.dumps({'progress': 0, 'status': f'Error: {str(e)}', 'complete': False, 'error': True})}\n\n"

    return StreamingResponse(batch_evaluation_stream(), media_type="text/event-stream")

# ==================== RESULTS API ====================

@app.get("/results")
async def get_results():
    try:
        results = get_all_results()
        formatted = []
        for r in results:
            if r['roll_no'] in ['ERROR', ''] or r['subject_name'] in ['ERROR', '']:
                continue
            formatted.append({
                "id": f"{r['roll_no']}_{r['subject_name']}",
                "roll_no": r['roll_no'],
                "name": r['name'],
                "subject_code": r['subject_code'] or 'N/A',
                "subject_name": r['subject_name'],
                "marks_obtained": r['marks_obtained'],
                "max_marks": r['max_marks'],
                "image_paths": r['image_paths']
            })
        return formatted
    except Exception as e:
        print(f"Error fetching results: {e}")
        return []

@app.get("/result/{roll}")
async def get_result(roll: str, subject_name: str = None):
    try:
        result = get_result_by_roll(roll, subject_name)
        if not result:
            return JSONResponse({"error": "Not found"}, status_code=404)
        if isinstance(result, list):
            result = result[0]

        feedback_data = result.get('feedback', [])
        if isinstance(feedback_data, str):
            try:
                feedback_data = json.loads(feedback_data)
            except:
                feedback_data = []

        marks_obtained = result.get('marks_obtained', result.get('marks', 0))
        max_marks = result.get('max_marks', result.get('total_max', 0))

        return {
            "roll_no": result['roll_no'],
            "name": result['name'],
            "subject_code": result.get('subject_code', 'N/A'),
            "subject_name": result['subject_name'],
            "marks_obtained": marks_obtained,
            "max_marks": max_marks,
            "percentage": round((marks_obtained/max_marks)*100, 1) if max_marks > 0 else 0,
            "image_paths": result.get('image_paths', []),
            "feedback": feedback_data,
            "markings": result.get('markings', {})
        }
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JSONResponse({"error": str(e)}, status_code=500)

@app.delete("/result/{roll}")
async def delete_result_api(roll: str, subject_name: str = None):
    try:
        if not subject_name:
            return JSONResponse({"error": "Subject name required"}, status_code=400)

        success = delete_result(roll, subject_name)
        if success:
            return {"message": f"Deleted {roll} - {subject_name}"}
        return JSONResponse({"error": "Result not found"}, status_code=404)
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)

@app.delete("/results/clear-invalid")
async def clear_invalid():
    import sqlite3
    conn = sqlite3.connect('results.db')
    cursor = conn.cursor()
    cursor.execute("DELETE FROM results WHERE roll_no IN ('ERROR', '') OR subject_name IN ('ERROR', '')")
    deleted = cursor.rowcount
    conn.commit()
    conn.close()
    return {"message": f"Deleted {deleted} invalid entries"}

# ==================== DOWNLOAD API ====================

@app.get("/download/excel/all")
async def download_all_excel():
    try:
        results = get_all_results()
        results = [r for r in results if r['roll_no'] not in ['ERROR', ''] and r['subject_name'] not in ['ERROR', '']]

        if not results:
            return JSONResponse({"error": "No valid results found"}, status_code=404)

        wb = Workbook()
        ws = wb.active
        ws.title = "All Results"

        header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)

        headers = ['Roll No', 'Name', 'Subject Code', 'Subject Name', 'Marks Obtained', 'Max Marks', 'Percentage', 'Status']
        for idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=idx, value=header)
            cell.fill = header_fill
            cell.font = header_font

        for row_idx, r in enumerate(results, 2):
            pct = round((r['marks_obtained']/r['max_marks']*100), 1) if r['max_marks'] > 0 else 0
            status = "OK" if r['subject_code']!= 'NOT_FOUND' and r['subject_code']!= '' else "Code Missing"

            ws.cell(row=row_idx, column=1, value=r['roll_no'])
            ws.cell(row=row_idx, column=2, value=r['name'])
            ws.cell(row=row_idx, column=3, value=r['subject_code'] if r['subject_code']!= 'NOT_FOUND' else 'N/A')
            ws.cell(row=row_idx, column=4, value=r['subject_name'])
            ws.cell(row=row_idx, column=5, value=r['marks_obtained'])
            ws.cell(row=row_idx, column=6, value=r['max_marks'])
            ws.cell(row=row_idx, column=7, value=f"{pct}%")
            ws.cell(row=row_idx, column=8, value=status)

        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
            ws.column_dimensions[col].width = 18

        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        return StreamingResponse(
            excel_file,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=All_Results.xlsx"}
        )
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JSONResponse({"error": str(e)}, status_code=500)

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)